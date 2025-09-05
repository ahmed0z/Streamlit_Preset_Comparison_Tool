"""
Export and results management utilities
"""
import pandas as pd
import io
from typing import List, Dict, Any, Optional
from datetime import datetime
import config

class ExportManager:
    """Handles export of comparison results to various formats"""
    
    @staticmethod
    def create_results_excel(results_data: List[Dict[str, Any]], 
                           input_summary: Dict[str, Any] = None) -> bytes:
        """
        Create Excel file with comparison results
        
        Args:
            results_data: List of comparison results
            input_summary: Optional summary information
            
        Returns:
            bytes: Excel file content
        """
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main results sheet
            if results_data:
                results_df = pd.DataFrame(results_data)
                results_df.to_excel(writer, sheet_name='Comparison Results', index=False)
                
                # Format the results sheet
                ExportManager._format_results_sheet(writer, 'Comparison Results', results_df)
            
            # Summary sheet
            if input_summary:
                summary_df = ExportManager._create_summary_dataframe(input_summary, results_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Statistics sheet
            if results_data:
                stats_df = ExportManager._create_statistics_dataframe(results_data)
                stats_df.to_excel(writer, sheet_name='Statistics', index=False)
            
            # Processing details sheet
            details_df = ExportManager._create_processing_details()
            details_df.to_excel(writer, sheet_name='Processing Details', index=False)
        
        output.seek(0)
        return output.read()
    
    @staticmethod
    def _format_results_sheet(writer, sheet_name: str, df: pd.DataFrame):
        """Apply formatting to the results sheet"""
        try:
            from openpyxl.styles import PatternFill, Font, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            workbook = writer.book
            worksheet = workbook[sheet_name]
            
            # Define colors
            exact_match_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
            similar_match_fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
            not_found_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
            header_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            
            # Header formatting
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            
            # Apply conditional formatting based on status
            status_col_idx = None
            for idx, col in enumerate(df.columns, 1):
                if col.lower() == 'status':
                    status_col_idx = idx
                    break
            
            if status_col_idx:
                for row_num in range(2, len(df) + 2):
                    status_cell = worksheet.cell(row=row_num, column=status_col_idx)
                    status_value = status_cell.value
                    
                    # Apply row coloring based on status
                    if status_value == config.MatchStatus.EXACT_MATCH:
                        fill = exact_match_fill
                    elif status_value == config.MatchStatus.PARTIAL_MATCH:
                        fill = similar_match_fill
                    elif status_value == config.MatchStatus.NOT_FOUND:
                        fill = not_found_fill
                    else:
                        continue
                    
                    # Apply fill to entire row
                    for col_num in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_num, column=col_num).fill = fill
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except ImportError:
            # If openpyxl styling is not available, continue without formatting
            pass
        except Exception as e:
            # If formatting fails, continue without it
            pass
    
    @staticmethod
    def _create_summary_dataframe(input_summary: Dict[str, Any], 
                                 results_data: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create summary information dataframe"""
        
        # Calculate statistics
        total_inputs = len(results_data) if results_data else 0
        exact_matches = sum(1 for r in results_data if r.get('Status') == config.MatchStatus.EXACT_MATCH)
        similar_matches = sum(1 for r in results_data if r.get('Status') == config.MatchStatus.PARTIAL_MATCH)
        not_found = sum(1 for r in results_data if r.get('Status') == config.MatchStatus.NOT_FOUND)
        
        avg_similarity = 0
        if results_data:
            similarities = [r.get('Similarity %', 0) for r in results_data if r.get('Similarity %', 0) > 0]
            avg_similarity = sum(similarities) / len(similarities) if similarities else 0
        
        summary_data = {
            'Metric': [
                'Processing Date',
                'Total Input Values',
                'Exact Matches',
                'Similar Matches', 
                'Not Found',
                'Match Rate (%)',
                'Average Similarity (%)',
                'Input File',
                'Preset Database'
            ],
            'Value': [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                total_inputs,
                exact_matches,
                similar_matches,
                not_found,
                round(((exact_matches + similar_matches) / total_inputs * 100) if total_inputs > 0 else 0, 1),
                round(avg_similarity, 1),
                input_summary.get('filename', 'Unknown') if input_summary else 'Unknown',
                'preset_database.xlsx'
            ]
        }
        
        return pd.DataFrame(summary_data)
    
    @staticmethod
    def _create_statistics_dataframe(results_data: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create detailed statistics dataframe"""
        if not results_data:
            return pd.DataFrame({'Statistic': ['No data'], 'Value': ['N/A']})
        
        # Similarity distribution
        similarities = [r.get('Similarity %', 0) for r in results_data]
        
        similarity_ranges = {
            '90-100%': sum(1 for s in similarities if s >= 90),
            '80-89%': sum(1 for s in similarities if 80 <= s < 90),
            '75-79%': sum(1 for s in similarities if 75 <= s < 80),
            'Below 75%': sum(1 for s in similarities if s < 75)
        }
        
        # Status distribution
        status_counts = {}
        for result in results_data:
            status = result.get('Status', 'Unknown')
            status_counts[status] = status_counts.get(status, 0) + 1
        
        # Combine statistics
        stats_data = []
        
        # Add similarity range statistics
        stats_data.append({'Category': 'Similarity Ranges', 'Metric': '', 'Count': '', 'Percentage': ''})
        for range_name, count in similarity_ranges.items():
            percentage = (count / len(results_data)) * 100 if results_data else 0
            stats_data.append({
                'Category': '',
                'Metric': range_name,
                'Count': count,
                'Percentage': f"{percentage:.1f}%"
            })
        
        # Add status statistics
        stats_data.append({'Category': 'Match Status', 'Metric': '', 'Count': '', 'Percentage': ''})
        for status, count in status_counts.items():
            percentage = (count / len(results_data)) * 100 if results_data else 0
            stats_data.append({
                'Category': '',
                'Metric': status,
                'Count': count,
                'Percentage': f"{percentage:.1f}%"
            })
        
        return pd.DataFrame(stats_data)
    
    @staticmethod
    def _create_processing_details() -> pd.DataFrame:
        """Create processing details and configuration"""
        details_data = {
            'Setting': [
                'Matching Threshold',
                'Exact Match Threshold',
                'Max Results Per Input',
                'Fuzzy Matching Algorithms',
                'Processing Date',
                'Tool Version'
            ],
            'Value': [
                f"{config.MATCHING_THRESHOLD * 100}%",
                f"{config.EXACT_MATCH_THRESHOLD * 100}%",
                config.MAX_RESULTS_PER_INPUT,
                'Levenshtein, Jaro-Winkler, Token Sort, Token Set, WRatio',
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                config.VERSION
            ],
            'Description': [
                'Minimum similarity score required for a match',
                'Threshold for considering a match as "exact"',
                'Maximum number of matches returned per input value',
                'Fuzzy string matching algorithms used',
                'When this analysis was performed',
                'Version of the comparison tool'
            ]
        }
        
        return pd.DataFrame(details_data)
    
    @staticmethod
    def create_csv_export(results_data: List[Dict[str, Any]]) -> str:
        """
        Create CSV export of results
        
        Args:
            results_data: List of comparison results
            
        Returns:
            str: CSV content
        """
        if not results_data:
            return "No data to export"
        
        df = pd.DataFrame(results_data)
        return df.to_csv(index=False)
    
    @staticmethod
    def create_filtered_export(results_data: List[Dict[str, Any]], 
                              filter_criteria: Dict[str, Any]) -> bytes:
        """
        Create filtered export based on criteria
        
        Args:
            results_data: List of comparison results
            filter_criteria: Filter conditions
            
        Returns:
            bytes: Filtered Excel file content
        """
        if not results_data:
            return ExportManager.create_results_excel([])
        
        df = pd.DataFrame(results_data)
        
        # Apply filters
        if filter_criteria.get('status'):
            df = df[df['Status'] == filter_criteria['status']]
        
        if filter_criteria.get('min_similarity'):
            df = df[df['Similarity %'] >= filter_criteria['min_similarity']]
        
        if filter_criteria.get('max_similarity'):
            df = df[df['Similarity %'] <= filter_criteria['max_similarity']]
        
        if filter_criteria.get('search_term'):
            search_term = filter_criteria['search_term'].lower()
            mask = (df['Original Input'].str.lower().str.contains(search_term, na=False) |
                   df['Matched Preset Value'].str.lower().str.contains(search_term, na=False))
            df = df[mask]
        
        return ExportManager.create_results_excel(df.to_dict('records'))

class ResultsAnalyzer:
    """Analyzes comparison results for insights"""
    
    @staticmethod
    def analyze_results(results_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Analyze comparison results for patterns and insights
        
        Args:
            results_data: List of comparison results
            
        Returns:
            Dict: Analysis insights
        """
        if not results_data:
            return {'message': 'No results to analyze'}
        
        df = pd.DataFrame(results_data)
        
        analysis = {
            'total_inputs': len(df),
            'unique_inputs': df['Original Input'].nunique(),
            'status_distribution': df['Status'].value_counts().to_dict(),
            'similarity_stats': {
                'mean': df['Similarity %'].mean(),
                'median': df['Similarity %'].median(),
                'std': df['Similarity %'].std(),
                'min': df['Similarity %'].min(),
                'max': df['Similarity %'].max()
            }
        }
        
        # Find common patterns
        analysis['insights'] = ResultsAnalyzer._generate_insights(df)
        
        return analysis
    
    @staticmethod
    def _generate_insights(df: pd.DataFrame) -> List[str]:
        """Generate insights from the results"""
        insights = []
        
        total = len(df)
        exact_matches = len(df[df['Status'] == config.MatchStatus.EXACT_MATCH])
        similar_matches = len(df[df['Status'] == config.MatchStatus.PARTIAL_MATCH])
        not_found = len(df[df['Status'] == config.MatchStatus.NOT_FOUND])
        
        # Match rate insights
        match_rate = (exact_matches + similar_matches) / total * 100
        if match_rate >= 90:
            insights.append(f"Excellent match rate: {match_rate:.1f}% of inputs found matches")
        elif match_rate >= 70:
            insights.append(f"Good match rate: {match_rate:.1f}% of inputs found matches")
        else:
            insights.append(f"Consider reviewing input data quality - {match_rate:.1f}% match rate")
        
        # Exact match insights
        exact_rate = exact_matches / total * 100
        if exact_rate >= 50:
            insights.append(f"High data consistency: {exact_rate:.1f}% exact matches")
        elif exact_rate < 20:
            insights.append("Many values require normalization - consider standardizing input formats")
        
        # Not found insights
        if not_found > 0:
            not_found_rate = not_found / total * 100
            if not_found_rate > 30:
                insights.append(f"Consider expanding preset database - {not_found_rate:.1f}% values not found")
        
        # Similarity insights
        avg_similarity = df['Similarity %'].mean()
        if avg_similarity >= 85:
            insights.append("High overall similarity scores suggest good data alignment")
        elif avg_similarity < 70:
            insights.append("Low similarity scores may indicate data format mismatches")
        
        return insights